from urllib.request import urlopen
from urllib.error import HTTPError
from nltk.tokenize import word_tokenize, sent_tokenize, RegexpTokenizer
import os
from bs4 import BeautifulSoup
import openpyxl
import re
import pandas as pd

input_file = openpyxl.load_workbook("Input.xlsx")
column_used = input_file.active

output_file = openpyxl.load_workbook("Output Data Structure.xlsx")
ws = output_file['Sheet1']

file1 = "StopWords\StopWords_Auditor.txt"
file2 = "StopWords\StopWords_Currencies.txt"
file3 = "StopWords\StopWords_DatesandNumbers.txt"
file4 = "StopWords\StopWords_Generic.txt"
file5 = "StopWords\StopWords_GenericLong.txt"
file6 = "StopWords\StopWords_Geographic.txt"
file7 = "StopWords\StopWords_Names.txt"
file8 = "MasterDictionary/negative_words.txt"
file9 = "MasterDictionary/positive_words.txt"

with open(file1,"r") as f1:
        auditor = [word.strip() for word in f1.readlines()]
with open(file2,"r") as f2:
    currencies = [word.strip() for word in f2.readlines()]
with open(file3,"r") as f3:
    datesAndNumbers = [word.strip() for word in f3.readlines()]
with open(file4,"r") as f4:
    generic = [word.strip() for word in f4.readlines()]
with open(file5,"r") as f5:
    genericLong = [word.strip() for word in f5.readlines()]
with open(file6,"r") as f6:
    geographic = [word.strip() for word in f6.readlines()]
with open(file7,"r") as f7:
    names = [word.strip() for word in f7.readlines()]
with open(file8,"r") as f8:
    negative = [word.strip() for word in f8.readlines()]
with open(file9,"r") as f9:
    positive = [word.strip()   for word in f9.readlines()]

##  All the words in the stopWords folder      
all_words = auditor + currencies + datesAndNumbers + generic + genericLong + geographic + names

## stopWords() will remove all the stopWords from the given text and will return a text that does not have stopWords
def stopWords(title):
    title_split = word_tokenize(title)
    analyse_words = [word for word in title_split if word not in all_words]

    analysis_words = " ".join(analyse_words)
    return analysis_words

## Column numbers to which the parameters have to be written
column_numbers = ['C','D','E','F','G','H','I','J','K','L','M','N','O']

## Creates a file for each url and contains only the title and actual content 
def create_files(url, url_id):
        try:
            with urlopen(url) as response:
                html_content = response.read()

                soup = BeautifulSoup(html_content, "html.parser")

                title = soup.find("h1")
                title = title.text if title else "No title found"

                # Extract the content
                content_element = soup.find_all("p")
                content = "\n ".join(paragraph.get_text() for paragraph in content_element)

                content = content.replace("Ranking customer behaviours for business strategy", "")
                content = content.replace("Algorithmic trading for multiple commodities markets, like Forex, Metals, Energy, etc.", "")
                content = content.replace("Trading Bot for FOREX", "")
                content = content.replace("Python model for the analysis of sector-specific stock ETFs for investment purposes", "")
                content = content.replace("Playstore & Appstore to Google Analytics (GA) or Firebase to Google Data Studio Mobile App KPI Dashboard", "")
                content = content.replace("Google Local Service Ads LSA API To Google BigQuery to Google Data Studio", "")
                content = content.replace("AI Conversational Bot using RASA", "")
                content = content.replace("Recommendation System Architecture", "")
                content = content.replace("Rise of telemedicine and its Impact on Livelihood by 2040", "")
                content = content.replace("Rise of e-health and its impact on humans by the year 2030", "")
                content = content.replace("Rise of e-health and its impact on humans by the year 2030", "")
                content = content.replace("Rise of telemedicine and its Impact on Livelihood by 2040", "")
                content = content.replace("AI/ML and Predictive Modeling", "")
                content = content.replace("Solution for Contact Centre Problems", "")
                content = content.replace("How to Setup Custom Domain for Google App Engine Application?", "")
                content = content.replace("Code Review Checklist", "")
                
                content = content.replace("Contact us: hello@blackcoffer.com ","")
                content = content.replace("© All Right Reserved, Blackcoffer(OPC) Pvt. Ltd","")

                filename = str(int(url_id)) + ".txt"

                with open(filename, "w", encoding="utf-8") as file:
                        file.write(title)
                        file.write("\n")
                        file.write(content)
##               if file status  code = 404 print error message 
        except HTTPError as e:
            if e.code == 404:
                print("Error")
            else:
                print("Error")


def segment_analysis(words):
    length = len(words)
    positive_words = 0
    negative_words  = 0
    
    for word in word_tokenize(words):
        if word in negative:
            negative_words -= 1
        if word in positive:
            positive_words += 1
            
    negative_words *= -1   
    polarity = (positive_words - negative_words)/ ((positive_words + negative_words) + 0.000001)
    subjectivity = (positive_words + negative_words)/(length + 0.000001)

    return negative_words , positive_words, polarity, subjectivity


def analysis_of_readability(text, complex_count, syllable_count):
        no_of_sentences = len(sent_tokenize(text))
        no_of_words = len(text.split(" "))
        words = text.split(" ")

        total_length = sum(len(word) for word in words)

        avg_sentence_length = no_of_words / no_of_sentences
        percent_complex = complex_count / no_of_words
        fog_index = 0.4 * (avg_sentence_length + percent_complex)

        avg_word_length = total_length/ no_of_words
        avg_syllable =  syllable_count / no_of_words
        return avg_sentence_length, percent_complex, fog_index, avg_syllable, avg_word_length

        
def syllable(word):
    word = word.lower()

    # exception_add - need extra syllables
    # exception_del - need less syllables

    exception_add = ['serious','crucial']
    exception_del = ['fortunately','unfortunately']

## co_one -     starts with co and has one syllable next to it
## co_two -     starts with co and has two syllable next to it
    co_one = ['cool','coach','coat','coal','count','coin','coarse','coup','coif','cook','coign','coiffe','coof','court']
    co_two = ['coapt','coed','coinci']

    pre_one = ['preach']

    syllable_count = 0 #added syllable number
    discarded_syllable = 0 #discarded syllable number

    #1) if letters < 3 : return 1
    if len(word) <= 3 :
        syllable_count = 1
        return syllable_count

    #2) if doesn't end with "ted" or "tes" or "ses" or "ied" or "ies", discard "es" and "ed" at the end.
    # if it has only 1 vowel or 1 set of consecutive vowels, discard. (like "speed", "fled" etc.)

    if word[-2:] == "es" or word[-2:] == "ed" :
        doubleAndtripple_1 = len(re.findall(r'[eaoui][eaoui]',word))
        if doubleAndtripple_1 > 1 or len(re.findall(r'[eaoui][^eaoui]',word)) > 1 :
            if word[-3:] == "ted" or word[-3:] == "tes" or word[-3:] == "ses" or word[-3:] == "ied" or word[-3:] == "ies" :
                pass
            else :
                discarded_syllable+=1

    #3) discard trailing "e", except where ending is "le"  

    le_except = ['whole','mobile','pole','male','female','hale','pale','tale','sale','aisle','whale','while']

    if word[-1:] == "e" :
        if word[-2:] == "le" and word not in le_except :
            pass

        else :
            discarded_syllable+=1

    #4) check if consecutive vowels exists, triplets or pairs, count them as one.

    doubleAndtripple = len(re.findall(r'[eaoui][eaoui]',word))
    tripple = len(re.findall(r'[eaoui][eaoui][eaoui]',word))
    discarded_syllable+=doubleAndtripple + tripple

    #5) count remaining vowels in word.
    numVowels = len(re.findall(r'[eaoui]',word))

    #6) add one if starts with "mc"
    if word[:2] == "mc" :
        syllable_count+=1

    #7) add one if ends with "y" but is not surrouned by vowel
    if word[-1:] == "y" and word[-2] not in "aeoui" :
        syllable_count +=1

    #8) add one if "y" is surrounded by non-vowels and is not in the last word.

    for i,j in enumerate(word) :
        if j == "y" :
            if (i != 0) and (i != len(word)-1) :
                if word[i-1] not in "aeoui" and word[i+1] not in "aeoui" :
                    syllable_count+=1

    #9) if starts with "tri-" or "bi-" and is followed by a vowel, add one.

    if word[:3] == "tri" and word[3] in "aeoui" :
        syllable_count+=1

    if word[:2] == "bi" and word[2] in "aeoui" :
        syllable_count+=1

    #10) if ends with "-ian", should be counted as two syllables, except for "-tian" and "-cian"

    if word[-3:] == "ian" : 
    #and (word[-4:] != "cian" or word[-4:] != "tian") :
        if word[-4:] == "cian" or word[-4:] == "tian" :
            pass
        else :
            syllable_count+=1

    #11) if starts with "co-" and is followed by a vowel, check if exists in the double syllable dictionary, if not, check if in single dictionary and act accordingly.

    if word[:2] == "co" and word[2] in 'eaoui' :

        if word[:4] in co_two or word[:5] in co_two or word[:6] in co_two :
            syllable_count+=1
        elif word[:4] in co_one or word[:5] in co_one or word[:6] in co_one :
            pass
        else :
            syllable_count+=1

    #12) if starts with "pre-" and is followed by a vowel, check if exists in the double syllable dictionary, if not, check if in single dictionary and act accordingly.

    if word[:3] == "pre" and word[3] in 'eaoui' :
        if word[:6] in pre_one :
            pass
        else :
            syllable_count+=1

    #13) check for "-n't" and cross match with dictionary to add syllable.

    negative = ["doesn't", "isn't", "shouldn't", "couldn't","wouldn't"]

    if word[-3:] == "n't" :
        if word in negative :
            syllable_count+=1
        else :
            pass   

    #14) Handling the exceptional words.

    if word in exception_del :
        discarded_syllable+=1

    if word in exception_add :
        syllable_count+=1     

    # calculate the output
    syllables = numVowels - discarded_syllable + syllable_count
    return syllables 

def complex_count(words):
        complex_word_count = 0
        syllable_count = 0
        for word in word_tokenize(words):
                if  syllable(word) > 2:
                        complex_word_count += 1
                        syllable_count += syllable(word)
        return complex_word_count, syllable_count

def averages(words):
        tokenizer = RegexpTokenizer(r'\w+')
        tokens = tokenizer.tokenize(words)
        
        us_count = 0

        personal_pronouns = ['I', 'you', 'he', 'she', 'it', 'we', 'they', 'me', 'him', 'her', 'us', 'them', 'myself',
                             'yourself', 'himself', 'herself', 'itself', 'ourselves', 'yourselves', 'themselves']

        
        if tokens == 'US':
                us_count += 1
                
        word_count = len(tokens) 
        
        personal_pronoun_count = sum(1 for word in tokens if word.lower() in personal_pronouns)

        return word_count - us_count, personal_pronoun_count

i=3
row_index = 2
column_letters = 'A'

for row in column_used.iter_rows(min_row =2, values_only=True):
        url = row[1]
        url_id = row[0]
        complex_word_count  = 0
        create_files(url, url_id)

        filename = str(int(url_id)) + ".txt"
        
        if os.path.isfile(filename) :
                with open(filename, "r", encoding='utf-8') as file:
                        text = file.readlines()

                text = ' '.join(text)
                words = stopWords(text)
                
                negative_words , positive_words, polarity, subjectivity = segment_analysis(words)

                complex_word_count , syllable_count = complex_count(words)
                
                avg_sentence_length, percent_complex, fog_index, avg_syllable, avg_word_length = analysis_of_readability(words, complex_word_count, syllable_count)

                word_count, pronoun_count  = averages(words)


                data = [positive_words, negative_words , polarity, subjectivity, avg_sentence_length, percent_complex, fog_index, avg_sentence_length, complex_word_count , word_count, avg_syllable, avg_word_length, pronoun_count]

## Writing back to the xlsx file
                excluded_row = [44,57,144]
                cell_value = ws[f'{column_letters}{row_index}'].value
                for index, column_letter in enumerate(column_numbers):
                        if cell_value not in excluded_row:
                                cell = ws[f'{column_letter}{row_index}']
                                cell.value = data[index]
                        else:
                                cell = ws[f'{column_letter}{row_index}']
                                cell.value = 0
                    
                row_index += 1
                        
                output_file.save('Output Data Structure.xlsx')

                print("finished")
        else:
                continue
